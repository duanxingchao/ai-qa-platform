"""
答案生成管理服务
"""

import logging
import pandas as pd
import tempfile
import os
from datetime import datetime, timedelta
from typing import Dict, List, Any, Optional, Tuple
from sqlalchemy import and_, or_
from app.models.question import Question
from app.models.answer import Answer
from app.utils.database import db


class AnswerGenerationService:
    """答案生成管理服务"""
    
    def __init__(self):
        self.logger = logging.getLogger(__name__)

    def _check_all_three_answers_exist(self, question_business_id: str) -> bool:
        """检查问题是否有完整的三个AI答案（yoyo, doubao, xiaotian）"""
        try:
            required_types = {'yoyo', 'doubao', 'xiaotian'}

            # 查询该问题的所有答案类型
            existing_answer_types = db.session.query(Answer.assistant_type).filter(
                and_(
                    Answer.question_business_id == question_business_id,
                    Answer.answer_text.isnot(None),
                    Answer.answer_text != ''
                )
            ).all()

            # 转换为集合
            existing_types = {answer_type[0] for answer_type in existing_answer_types}

            # 检查是否包含所有必需的答案类型
            is_complete = required_types.issubset(existing_types)

            self.logger.debug(f"问题 {question_business_id} 答案完整性检查: 需要{required_types}, 现有{existing_types}, 完整={is_complete}")

            return is_complete

        except Exception as e:
            self.logger.error(f"检查问题 {question_business_id} 答案完整性时出错: {str(e)}")
            return False

    def _trigger_scoring_if_ready(self) -> bool:
        """检查是否可以触发评分阶段，如果可以则触发"""
        try:
            # 检查是否还有待处理的问题
            pending_count = self.get_export_questions_count()

            if pending_count == 0:
                # 没有待处理问题，可以触发评分阶段
                self.logger.info("手动答案导入完成，所有问题都已有完整答案，准备触发评分阶段")

                # 动态导入以避免循环导入
                from flask import current_app
                from app.services.scheduler_service import scheduler_service, WorkflowPhase

                # 检查当前是否为手动模式
                answer_generation_mode = current_app.config.get('ANSWER_GENERATION_MODE', 'api')
                if answer_generation_mode == 'manual':
                    # 触发评分阶段
                    result = scheduler_service.execute_workflow_phase(
                        current_app._get_current_object(),
                        WorkflowPhase.SCORING
                    )

                    if result.get('success', False):
                        self.logger.info("评分阶段已成功触发")
                        return True
                    else:
                        self.logger.warning(f"评分阶段触发失败: {result.get('message', '未知错误')}")
                        return False
                else:
                    self.logger.info("当前为API模式，不需要手动触发评分阶段")
                    return False
            else:
                self.logger.info(f"还有{pending_count}个问题待处理，暂不触发评分阶段")
                return False

        except Exception as e:
            self.logger.error(f"触发评分阶段时出错: {str(e)}")
            return False
    
    def get_export_questions_count(self) -> int:
        """获取待导出问题数量 - 必须有yoyo答案但缺少doubao/xiaotian答案"""
        try:
            # 查询已分类且有yoyo答案但没有豆包和小天答案的问题
            count = db.session.query(Question).filter(
                and_(
                    Question.processing_status == 'classified',
                    Question.classification.isnot(None),
                    Question.classification != '',
                    # 必须有yoyo答案
                    db.session.query(Answer).filter(
                        and_(
                            Answer.question_business_id == Question.business_id,
                            Answer.assistant_type == 'yoyo',
                            Answer.answer_text.isnot(None),
                            Answer.answer_text != ''
                        )
                    ).exists(),
                    # 确保没有豆包答案
                    ~db.session.query(Answer).filter(
                        and_(
                            Answer.question_business_id == Question.business_id,
                            Answer.assistant_type == 'doubao'
                        )
                    ).exists(),
                    # 确保没有小天答案
                    ~db.session.query(Answer).filter(
                        and_(
                            Answer.question_business_id == Question.business_id,
                            Answer.assistant_type == 'xiaotian'
                        )
                    ).exists()
                )
            ).count()

            self.logger.info(f"待导出问题数量: {count} (已分类+有yoyo答案+缺少doubao/xiaotian答案)")
            return count

        except Exception as e:
            self.logger.error(f"获取待导出问题数量时出错: {str(e)}")
            raise
    
    def export_questions_to_excel(
        self, 
        time_range: Optional[str] = None,
        batch_size: Optional[int] = None
    ) -> Tuple[str, str]:
        """
        导出问题到Excel文件
        
        Args:
            time_range: 时间范围筛选 ('week', 'month', 'all')
            batch_size: 批次大小限制
            
        Returns:
            Tuple[str, str]: (文件路径, 文件名)
        """
        try:
            # 构建查询条件 - 已分类且有yoyo答案但缺少doubao/xiaotian答案的问题
            query = db.session.query(Question).filter(
                and_(
                    Question.processing_status == 'classified',
                    Question.classification.isnot(None),
                    Question.classification != '',
                    # 必须有yoyo答案
                    db.session.query(Answer).filter(
                        and_(
                            Answer.question_business_id == Question.business_id,
                            Answer.assistant_type == 'yoyo',
                            Answer.answer_text.isnot(None),
                            Answer.answer_text != ''
                        )
                    ).exists(),
                    # 确保没有豆包答案
                    ~db.session.query(Answer).filter(
                        and_(
                            Answer.question_business_id == Question.business_id,
                            Answer.assistant_type == 'doubao'
                        )
                    ).exists(),
                    # 确保没有小天答案
                    ~db.session.query(Answer).filter(
                        and_(
                            Answer.question_business_id == Question.business_id,
                            Answer.assistant_type == 'xiaotian'
                        )
                    ).exists()
                )
            )
            
            # 添加时间范围筛选
            if time_range == 'week':
                week_ago = datetime.utcnow() - timedelta(days=7)
                query = query.filter(Question.created_at >= week_ago)
            elif time_range == 'month':
                month_ago = datetime.utcnow() - timedelta(days=30)
                query = query.filter(Question.created_at >= month_ago)

            # 按创建时间排序（必须在limit之前）
            query = query.order_by(Question.created_at.desc())

            # 添加批次大小限制
            if batch_size and batch_size > 0:
                query = query.limit(batch_size)

            # 执行查询
            questions = query.all()
            
            if not questions:
                raise ValueError("没有找到待导出的问题")
            
            # 生成文件名
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f'questions_for_answer_generation_{timestamp}.xlsx'

            # 创建Excel文件
            from openpyxl import Workbook
            import tempfile

            wb = Workbook()
            ws = wb.active
            ws.title = 'Questions'

            # 写入表头
            headers = ['business_id', 'question', 'classification', 'doubao_answer', 'xiaotian_answer']
            for col_idx, header in enumerate(headers, 1):
                ws.cell(row=1, column=col_idx, value=header)

            # 写入数据
            for row_idx, question in enumerate(questions, 2):
                ws.cell(row=row_idx, column=1, value=question.business_id)
                ws.cell(row=row_idx, column=2, value=question.query)
                ws.cell(row=row_idx, column=3, value=question.classification)
                ws.cell(row=row_idx, column=4, value='')  # doubao_answer
                ws.cell(row=row_idx, column=5, value='')  # xiaotian_answer

            # 保存到临时文件
            temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
            temp_file.close()
            wb.save(temp_file.name)
            wb.close()

            self.logger.info(f"成功导出{len(questions)}个问题到Excel文件: {filename}")
            return temp_file.name, filename

        except Exception as e:
            self.logger.error(f"导出问题到Excel时出错: {str(e)}")
            raise

    def validate_import_file(self, file_path: str) -> Dict[str, Any]:
        """
        验证导入文件格式
        
        Args:
            file_path: Excel文件路径
            
        Returns:
            Dict: 验证结果
        """
        try:
            # 读取Excel文件
            df = pd.read_excel(file_path)
            
            # 检查必需的列
            required_columns = ['business_id', 'question', 'doubao_answer', 'xiaotian_answer']
            missing_columns = [col for col in required_columns if col not in df.columns]
            
            if missing_columns:
                return {
                    'valid': False,
                    'error': f"缺少必需的列: {', '.join(missing_columns)}",
                    'required_columns': required_columns,
                    'actual_columns': list(df.columns)
                }
            
            # 检查数据行数
            total_rows = len(df)
            if total_rows == 0:
                return {
                    'valid': False,
                    'error': "文件中没有数据行",
                    'total_rows': 0
                }
            
            # 检查空值情况
            empty_business_id = df['business_id'].isna().sum()
            empty_doubao_answer = df['doubao_answer'].isna().sum()
            empty_xiaotian_answer = df['xiaotian_answer'].isna().sum()
            
            return {
                'valid': True,
                'total_rows': total_rows,
                'columns': list(df.columns),
                'data_quality': {
                    'empty_business_id': int(empty_business_id),
                    'empty_doubao_answer': int(empty_doubao_answer),
                    'empty_xiaotian_answer': int(empty_xiaotian_answer)
                },
                'preview': df.head(5).to_dict('records')  # 预览前5行
            }
            
        except Exception as e:
            self.logger.error(f"验证导入文件时出错: {str(e)}")
            return {
                'valid': False,
                'error': f"文件读取失败: {str(e)}"
            }

    def import_answers_from_excel(self, file_path: str) -> Dict[str, Any]:
        """
        从Excel文件导入答案数据（支持容错处理）

        Args:
            file_path: Excel文件路径

        Returns:
            Dict: 导入结果报告
        """
        try:
            # 读取Excel文件
            df = pd.read_excel(file_path)

            # 检查必需的列
            required_columns = ['business_id', 'question', 'doubao_answer', 'xiaotian_answer']
            missing_columns = [col for col in required_columns if col not in df.columns]

            if missing_columns:
                raise ValueError(f"缺少必需的列: {', '.join(missing_columns)}")

            total_rows = len(df)
            success_count = 0
            failed_count = 0
            failed_items = []

            self.logger.info(f"开始导入{total_rows}行答案数据")

            # 逐行处理，确保单行错误不影响其他行
            for index, row in df.iterrows():
                row_number = index + 2  # Excel行号（从2开始，因为第1行是标题）

                try:
                    # 验证必需字段
                    business_id = str(row['business_id']).strip() if pd.notna(row['business_id']) else ''
                    doubao_answer = str(row['doubao_answer']).strip() if pd.notna(row['doubao_answer']) else ''
                    xiaotian_answer = str(row['xiaotian_answer']).strip() if pd.notna(row['xiaotian_answer']) else ''

                    # 检查business_id是否为空
                    if not business_id:
                        failed_items.append({
                            'row_number': row_number,
                            'business_id': business_id,
                            'error': 'business_id为空'
                        })
                        failed_count += 1
                        continue

                    # 检查答案是否为空
                    if not doubao_answer:
                        failed_items.append({
                            'row_number': row_number,
                            'business_id': business_id,
                            'error': '豆包答案为空'
                        })
                        failed_count += 1
                        continue

                    if not xiaotian_answer:
                        failed_items.append({
                            'row_number': row_number,
                            'business_id': business_id,
                            'error': '小天答案为空'
                        })
                        failed_count += 1
                        continue

                    # 查找对应的问题
                    question = db.session.query(Question).filter(
                        Question.business_id == business_id
                    ).first()

                    if not question:
                        failed_items.append({
                            'row_number': row_number,
                            'business_id': business_id,
                            'error': 'business_id不存在于数据库中'
                        })
                        failed_count += 1
                        continue

                    # 检查是否已有答案（避免重复导入）
                    existing_doubao = db.session.query(Answer).filter(
                        and_(
                            Answer.question_business_id == question.business_id,
                            Answer.assistant_type == 'doubao'
                        )
                    ).first()

                    existing_xiaotian = db.session.query(Answer).filter(
                        and_(
                            Answer.question_business_id == question.business_id,
                            Answer.assistant_type == 'xiaotian'
                        )
                    ).first()

                    if existing_doubao and existing_xiaotian:
                        failed_items.append({
                            'row_number': row_number,
                            'business_id': business_id,
                            'error': '该问题已有答案，跳过重复导入'
                        })
                        failed_count += 1
                        continue

                    # 创建或更新豆包答案
                    if not existing_doubao:
                        doubao_answer_obj = Answer(
                            question_business_id=question.business_id,
                            assistant_type='doubao',
                            answer_text=doubao_answer,
                            created_at=datetime.utcnow()
                        )
                        db.session.add(doubao_answer_obj)
                    else:
                        existing_doubao.answer_text = doubao_answer
                        existing_doubao.updated_at = datetime.utcnow()

                    # 创建或更新小天答案
                    if not existing_xiaotian:
                        xiaotian_answer_obj = Answer(
                            question_business_id=question.business_id,
                            assistant_type='xiaotian',
                            answer_text=xiaotian_answer,
                            created_at=datetime.utcnow()
                        )
                        db.session.add(xiaotian_answer_obj)
                    else:
                        existing_xiaotian.answer_text = xiaotian_answer
                        existing_xiaotian.updated_at = datetime.utcnow()

                    # 检查三个答案是否都完整，只有完整时才更新状态为answers_generated
                    if self._check_all_three_answers_exist(question.business_id):
                        question.processing_status = 'answers_generated'
                        question.updated_at = datetime.utcnow()
                        self.logger.info(f"问题 {question.business_id} 三个AI答案完整，状态更新为answers_generated")
                    else:
                        # 答案不完整，保持classified状态
                        question.updated_at = datetime.utcnow()
                        self.logger.warning(f"问题 {question.business_id} 答案不完整，保持classified状态")

                    # 提交当前行的更改
                    db.session.commit()
                    success_count += 1

                    if success_count % 10 == 0:
                        self.logger.info(f"已成功导入{success_count}条记录")

                except Exception as row_error:
                    # 回滚当前行的更改
                    db.session.rollback()

                    error_msg = str(row_error)
                    self.logger.error(f"导入第{row_number}行时出错: {error_msg}")

                    failed_items.append({
                        'row_number': row_number,
                        'business_id': business_id if 'business_id' in locals() else 'unknown',
                        'error': error_msg
                    })
                    failed_count += 1

            # 计算成功率
            success_rate = f"{(success_count / total_rows * 100):.1f}%" if total_rows > 0 else "0%"

            # 检查是否可以触发评分阶段
            scoring_triggered = False
            if success_count > 0:
                try:
                    scoring_triggered = self._trigger_scoring_if_ready()
                except Exception as e:
                    self.logger.error(f"触发评分阶段时出错: {str(e)}")

            # 生成导入结果报告
            import_result = {
                'success': True,
                'summary': {
                    'total_rows': total_rows,
                    'success_count': success_count,
                    'failed_count': failed_count,
                    'success_rate': success_rate
                },
                'failed_items': failed_items,
                'scoring_triggered': scoring_triggered,
                'message': f'导入完成，{success_count}条记录成功，{failed_count}条记录失败' +
                          ('，已自动触发评分阶段' if scoring_triggered else '')
            }

            self.logger.info(f"答案导入完成: 总计{total_rows}行，成功{success_count}行，失败{failed_count}行")

            return import_result

        except Exception as e:
            # 确保回滚所有未提交的更改
            db.session.rollback()
            self.logger.error(f"导入答案时出错: {str(e)}")
            raise

    def get_import_history(self, page: int = 1, per_page: int = 20) -> Dict[str, Any]:
        """
        获取导入历史记录

        Args:
            page: 页码
            per_page: 每页数量

        Returns:
            Dict: 分页的历史记录
        """
        try:
            # 这里可以扩展为从专门的导入历史表中获取记录
            # 目前返回基本的统计信息

            # 统计最近的答案创建情况
            recent_answers = db.session.query(Answer).filter(
                Answer.assistant_type.in_(['doubao', 'xiaotian'])
            ).order_by(Answer.created_at.desc()).limit(per_page * page).all()

            # 按创建时间分组统计
            history_items = []
            current_date = None
            daily_count = 0

            for answer in recent_answers:
                answer_date = answer.created_at.date()
                if current_date != answer_date:
                    if current_date is not None:
                        history_items.append({
                            'date': current_date.isoformat(),
                            'count': daily_count,
                            'type': 'import'
                        })
                    current_date = answer_date
                    daily_count = 1
                else:
                    daily_count += 1

            # 添加最后一天的记录
            if current_date is not None:
                history_items.append({
                    'date': current_date.isoformat(),
                    'count': daily_count,
                    'type': 'import'
                })

            return {
                'items': history_items,
                'pagination': {
                    'page': page,
                    'per_page': per_page,
                    'total': len(history_items)
                }
            }

        except Exception as e:
            self.logger.error(f"获取导入历史时出错: {str(e)}")
            raise
